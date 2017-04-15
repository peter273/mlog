# -*- coding: utf-8 -*-

import sys
from PySide import QtCore, QtGui

from MineLog import Equipment,ShiftFile,mload,\
        oeeEquipmentPlot,Moving_AveragePlot,get_tframe

import openpyxl 
from openpyxl.utils.dataframe import dataframe_to_rows

import os


#New Equipment Dialog
from mlogUI.neweqdialog import Ui_Dialog
class NewEqDialog(QtGui.QDialog,Ui_Dialog):
    def __init__(self,parent=None):
        QtGui.QDialog.__init__(self,parent)
        self.setupUi(self)
    def getValues(self):
        name=self.eqName_lineedit.text()
        eqtype= self.eqtype_combobox.currentText()
        return (str(name).strip(),str(eqtype))
class EqItemGui:
    def __init__(self,eq):
        self.eq=eq
        self.eq_gui=self.get_all_data()
    def get_eq_gui(self):
        i=self.eq
        eq_item=QtGui.QTreeWidgetItem(None,[i.Name])
        attributes=["OEE",
                "Availability",
                "Utilization",
                "Efficiency",]

        for ctr in range(len(i.Data)):
            Date=str(i.Data.iloc[ctr].Date.date())
            Shift=i.Data.iloc[ctr].Shift
            param=[str(getattr(i.Data.iloc[ctr],temp)) for temp in attributes]
            shift_item=QtGui.QTreeWidgetItem(eq_item,[Date+" Shift "+Shift,
                *param])
        return eq_item
    def get_all_data(self):
        data={'shiftly':None,
                'daily':None,
                'weekly':None,
                'monthly':None}
        for i in data:
            data[i]=self.get_eqgui_data(i)
        return data
    def get_eqgui_data(self,data_frame):
        if data_frame=="shiftly":
            return self.get_eq_gui()
        eq_item=QtGui.QTreeWidgetItem(None,["{0} ({1})".format(self.eq.Name,data_frame)])
        attributes=["OEE",
                "Availability",
                "Utilization",
                "Efficiency"]
        if len(self.eq.Data):
            Data=get_tframe(self.eq,data_frame)
            for ctr in range(len(Data)):
                Date=str(Data.iloc[ctr].name.date())
                param=[str(getattr(Data.iloc[ctr],temp)) for temp in attributes]
                shift_item=QtGui.QTreeWidgetItem(eq_item,[Date,
                    *param])
        return eq_item



from mlogUI import mwindow
class MainUI(mwindow.Ui_MainWindow):
    def setupUi(self,MainWindow):
        super(MainUI,self).setupUi(MainWindow)
        dirname=os.path.dirname(__file__)

        icon_path={
                'new':['plus.png',self.actionNewEquipment],
                'delete':['cancel-2.png',self.actionDelete],
                'addshift':['copy.png',self.actionAdd_Shift],
                # 'info':['information.png',self.info_toolbtn],
                'download':['download.png',self.actionExport],
                'plot':['fast-forward.png',self.actionPlot]}

        for i in icon_path:
            icon_path[i][0]=os.path.join(dirname,"mlogUI/flaticons/icons/png/"+icon_path[i][0])
            icon = QtGui.QIcon()
            icon.addPixmap(QtGui.QPixmap(icon_path[i][0]), QtGui.QIcon.Normal, QtGui.QIcon.Off)
            icon_path[i][1].setIcon(icon)
            # icon_path[i][1].setIconSize(QtCore.QSize(30, 30))

        self.shovel_listwidget.setSelectionMode(QtGui.QAbstractItemView.ExtendedSelection)
        self.shovel_listwidget.itemSelectionChanged.connect(self.itemSelectionChanged)
        self.truck_listwidget.setSelectionMode(QtGui.QAbstractItemView.ExtendedSelection)
        self.truck_listwidget.itemSelectionChanged.connect(self.itemSelectionChanged)

        self.ma_checkbox.stateChanged.connect(self.maspinner_changestate)
        self.ma_spinBox.setEnabled(False)
        self.get_initial_equipment_list()

        self.shift_rb.toggled.connect(lambda:self.bntstate(self.shift_rb))
        self.day_rb.toggled.connect(lambda:self.bntstate(self.day_rb))
        self.week_rb.toggled.connect(lambda:self.bntstate(self.week_rb))
        self.month_rb.toggled.connect(lambda:self.bntstate(self.month_rb))
        self.display_headers()

    def bntstate(self,b):
        if b.isChecked():
            self.itemSelectionChanged()

    # Display Headers
    def display_headers(self):
        header=QtGui.QTreeWidgetItem([
            "Name",
            "OEE",
            "Availability",
            "Utilization",
            "Efficiency",
            "Total Hours", ])
        self.data_listwidget.setHeaderItem(header)
        chosen_dataframe=self.chosen_dataframe()
        for lw in [self.shovel_listwidget,self.truck_listwidget]:
            for i in range(lw.count()):
                g=lw.item(i).data(QtCore.Qt.UserRole).eq_gui
                for tframe in g:
                    self.data_listwidget.addTopLevelItem(g[tframe])
                    self.data_listwidget.setItemHidden(g[tframe],True)

    # Change Moving Average Spinner state when Moving Average check box state changes
    def maspinner_changestate(self):
        if self.ma_checkbox.isChecked():
            self.ma_spinBox.setEnabled(True)
        else:
            self.ma_spinBox.setEnabled(False)

    def itemSelectionChanged(self):
        eq = self.chosen_equipment()
        chosen_dataframe=self.chosen_dataframe()
        x=['shiftly','daily','weekly','monthly']
        eqpment=[j.data(QtCore.Qt.UserRole).eq_gui[chosen_dataframe] for j in eq]
        # Hide all Items in data_listwidget
        for lw in [self.shovel_listwidget,self.truck_listwidget]:
            for i in range(lw.count()):
                for tframe in x:
                    g=lw.item(i).data(QtCore.Qt.UserRole).eq_gui[tframe]
                    self.data_listwidget.setItemHidden(g,True)
        # Show all item from chosen epment in data_listwidget
        for i in eqpment:
            self.data_listwidget.setItemHidden(i,False)

    # Return the Time Frame to be Plotted
    def chosen_dataframe(self):
        x={'shiftly':self.shift_rb,
            'daily':self.day_rb,
            'weekly':self.week_rb,
            'monthly':self.month_rb}
        for i in x:
            if x[i].isChecked():
                return i
    # Return a list of Chosen Equipment
    def chosen_equipment(self):
        #0-shovels,1-trucks
        eqlist=self.eq_tab.currentIndex()
        if eqlist==0: eq_listwidget=self.shovel_listwidget
        elif eqlist==1: eq_listwidget=self.truck_listwidget

        x = eq_listwidget.selectedItems()
        return x
    # Returns a list of chosen parameters
    def chosen_params(self):
        params=[]
        cbox=[self.availability_checkbox,
                self.utilization_checkbox,
                self.eff_checkbox,
                self.oee_checkbox]
        for i in cbox:
            if i.checkState():
                params.append(i.text())
        return params

    def get_initial_equipment_list(self):
        dirpath=os.path.dirname(__file__)
        filepath = os.path.join(dirpath,'Equipment')
        eq_filenames = [i for i in os.listdir(filepath) if i.endswith(".mlog")]
        eq_list=[]

        listwidget=self.shovel_listwidget
        for i in eq_filenames:
            eq=mload(os.path.join(filepath,i))

            if eq.Type=="Shovel": 
                listwidget=self.shovel_listwidget
            else:
                listwidget=self.truck_listwidget

            eq_item=QtGui.QListWidgetItem()
            eq_item.setText(eq.Name)
            eq_item.setData(QtCore.Qt.UserRole,EqItemGui(eq))

            listwidget.addItem(eq_item)
        listwidget.repaint()

class MineLog(QtGui.QMainWindow):
    def __init__(self,parent=None):
        super(MineLog,self).__init__(parent)
        # self.ui=Ui_MainWindow()
        self.ui=MainUI()
        self.ui.setupUi(self)
        self.ui.actionNewEquipment.triggered.connect(self.new_buttonClicked)
        self.ui.actionAdd_Shift.triggered.connect(self.func_addshift)
        self.ui.actionDelete.triggered.connect(self.func_deleq)
        self.ui.actionPlot.triggered.connect(self.plt_btn_clicked)
        self.ui.actionExport.triggered.connect(self.export_data)
        #TODO check if Equipment folder exists
    def export_data(self):
        eq = self.ui.chosen_equipment()
        eqpment = [i.data(QtCore.Qt.UserRole).eq for i in eq]
        data_frame=self.ui.chosen_dataframe()
        params=self.ui.chosen_params()

        moving_ave_on=self.ui.ma_checkbox.checkState()
        interval = (self.ui.ma_spinBox.value())
        if eqpment and params and not moving_ave_on:
            self.export_data_helper(
                oeeEquipmentPlot(eqpment,params,data_frame=data_frame,
                    export_data=True),
                eqpment,
                data_frame,)

        elif moving_ave_on and eqpment and params:
            self.export_data_helper(
                Moving_AveragePlot(eqpment,params,data_frame=data_frame,
                    interval=interval,export_data=True),
                eqpment,
                data_frame,)


        elif not eq:
            msg = QtGui.QMessageBox(self)
            msg.setText('No Equipment was chosen')
            msg.setWindowTitle("Warning")
            msg.show()
    def export_data_helper(self,data,eqpment,data_frame):
        filename,filext = QtGui.QFileDialog.getSaveFileName(self,
            "Save file","Untitled",".xlsx")
        if filename:
        
            wb=openpyxl.Workbook(write_only=True)
            ws={i:wb.create_sheet(i) for i in data}

            for i in data:
                for row in dataframe_to_rows(data[i],index=True,header=True):
                    ws[i].append(row)
            print('done saving',filename+filext)
            wb.save(filename+filext)


    # Plotting of Data 
    def plt_btn_clicked(self):
        eq = self.ui.chosen_equipment()
        eqpment = [i.data(QtCore.Qt.UserRole).eq for i in eq]
        data_frame=self.ui.chosen_dataframe()
        params=self.ui.chosen_params()

        moving_ave_on=self.ui.ma_checkbox.checkState()
        interval = (self.ui.ma_spinBox.value())
        if eqpment and params and not moving_ave_on:
            oeeEquipmentPlot(eqpment,params,data_frame=data_frame)
        elif moving_ave_on and eqpment and params:
            Moving_AveragePlot(eqpment,params,data_frame=data_frame,interval=interval)
        elif not eq:
            msg = QtGui.QMessageBox(self)
            msg.setText('No Equipment was chosen')
            msg.setWindowTitle("Warning")
            msg.show()

    # Delete Equipment Function
    def func_deleq(self):
        x=self.ui.chosen_equipment()
        if len(x)>0:
            msg = "Are you sure you want to delete chosen Equipment?"
            reply =QtGui.QMessageBox.question(self,'Message',msg,QtGui.QMessageBox.Yes,QtGui.QMessageBox.No)
            if reply == QtGui.QMessageBox.Yes:
                self.delete_equipment()
        else:
            msg = QtGui.QMessageBox(self)
            msg.setText('No Chosen Equipment')
            msg.setWindowTitle("Warning")
            msg.show()
    def delete_equipment(self):
        dirpath=os.path.dirname(__file__)
        filepath = os.path.join(dirpath,'Equipment')
        if not os.path.exists(filepath):
            os.mkdir(filepath)

        eqlist=self.ui.eq_tab.currentIndex()
        if eqlist==0: eq_listwidget=self.ui.shovel_listwidget
        elif eqlist==1: eq_listwidget=self.ui.truck_listwidget
        
        x = eq_listwidget.selectedItems()
        for i in x:
            eq_listwidget.takeItem(eq_listwidget.row(i))
            g = i.data(QtCore.Qt.UserRole).eq
            os.remove(os.path.join(filepath,g.Name+".mlog"))

    # Add Shift Function
    def func_addshift(self):
        # TODO add shift option to add from folder
        dirpath=os.path.dirname(__file__)
        filepath = os.path.join(dirpath,'Equipment')
        eq=self.ui.chosen_equipment()
        if len(eq)==1:
            fname,fx = QtGui.QFileDialog.getOpenFileNames(self,'Add Shift File')
            g=eq[0].data(QtCore.Qt.UserRole)
            notadded=[]
            for sfile in fname:
                if g.eq.AddFile(sfile): 
                    print('{0} is added'.format(sfile))
                else:
                    notadded.append(os.path.basename(sfile))
                    print("Error in Adding {0}".format(sfile))
            if notadded:
                msg = "Error in Adding {0} out of {1} file/s:\n".format(len(notadded),len(fname))
                msg += "Do you wish to continue adding files that can be Added?"
                reply = QtGui.QMessageBox.question(self,'Message',msg,QtGui.QMessageBox.Yes,QtGui.QMessageBox.No)
                if reply == QtGui.QMessageBox.Yes:
                    g.eq.update()
                    g.eq.save(filepath)
            else:
                g.eq.update()
                g.eq.save(filepath)

                data_listwidget=self.ui.data_listwidget
                chosen_dataframe=self.ui.chosen_dataframe()
                for i in g.eq_gui:
                    print(i,g.eq_gui[i])

                temp = g.get_all_data()
                
                for i in g.eq_gui:
                    index_item=data_listwidget.indexOfTopLevelItem(g.eq_gui[i])
                    if index_item!=-1:
                        data_listwidget.takeTopLevelItem(index_item)
                        data_listwidget.insertTopLevelItem(index_item,temp[i])
                    else:
                        data_listwidget.addTopLevelItem(temp[i])
                        data_listwidget.setItemHidden(temp[i],True)

                g.eq_gui=temp
                self.ui.data_listwidget.setItemHidden(g.eq_gui[chosen_dataframe],False)
                self.ui.data_listwidget.repaint()

                # eqlist=self.ui.eq_tab.currentIndex()
                # if eqlist==0: eq_listwidget=self.ui.shovel_listwidget
                # elif eqlist==1: eq_listwidget=self.ui.truck_listwidget
                
                # x = eq_listwidget.selectedItems()
                # for i in x:
                #     eq_listwidget.takeItem(eq_listwidget.row(i))

                # ----------------
                ## g=lw.item(i).data(QtCore.Qt.UserRole).eq_gui

                # new_eq_gui=EqItemGui(g.eq)
                # eq_item=QtGui.QListWidgetItem()
                # eq_item.setText(name)
                # eq_item.setData(QtCore.Qt.UserRole,new_eq_gui)

                # if eq_type=="Shovel": listwidget=self.ui.shovel_listwidget
                # elif eq_type=="Truck": listwidget=self.ui.truck_listwidget

                # listwidget.addItem(eq_item)
                # listwidget.repaint()


        else:
            msg = QtGui.QMessageBox(self)
            msg.setText('Choose only one Equipment')
            msg.setWindowTitle("Warning")
            msg.show()
    # Opens Dialog for Creating New Equipment
    def new_buttonClicked(self):
        dialog = NewEqDialog(self)
        if dialog.exec_():
            eq_name,eq_type=dialog.getValues()
            if eq_name=="":
                msg = QtGui.QMessageBox(dialog)
                msg.setText('Equipment name cannot be empty')
                msg.setWindowTitle("Warning")
                msg.show()
            else: 
                self.create_equipment(eq_name,eq_type)
    # Function for creating New Equipment
    def create_equipment(self,name,eq_type):
        #TODO Check if Equipment Name Already exist
        import os
        dirpath= os.path.dirname(__file__)
        filepath= os.path.join(dirpath,'Equipment')
        if not os.path.exists(filepath):
            os.mkdir(filepath)
        new_eq=Equipment(name,eq_type)
        new_eq.save(filepath)
        new_eq_gui=EqItemGui(new_eq)

        eq_item=QtGui.QListWidgetItem()
        eq_item.setText(name)
        eq_item.setData(QtCore.Qt.UserRole,new_eq_gui)

        if eq_type=="Shovel": listwidget=self.ui.shovel_listwidget
        elif eq_type=="Truck": listwidget=self.ui.truck_listwidget

        listwidget.addItem(eq_item)
        listwidget.repaint()



if __name__=="__main__":
    app=QtGui.QApplication(sys.argv)
    win =MineLog()
    win.show()
    sys.exit(app.exec_())
